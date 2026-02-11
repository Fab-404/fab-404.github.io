import openpyxl
from openpyxl.utils import get_column_letter
import json
import os
import datetime

# Configuration
EXCEL_FILE = "TBO GLOBAL.xlsm"
OUTPUT_JS_FILE = "tbo-global-proto/js/demo-data.js"

class DateTimeEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, (datetime.date, datetime.datetime)):
            return obj.strftime('%d/%m/%Y')
        return super().default(obj)

def get_rgb_color(color_obj):
    """Convertit une couleur openpyxl en CSS hex."""
    if not color_obj:
        return None
    try:
        # Si type ThemeColor, .rgb peut ne pas être directement accessible ou valide sans conversion
        # On tente l'accès direct, mais on filtre les valeurs absurdes
        if hasattr(color_obj, 'rgb'):
            argb = str(color_obj.rgb)
            # Vérifier si c'est un hex valide (parfois openpyxl renvoie des erreurs en string)
            if len(argb) >= 6 and all(c in '0123456789ABCDEFabcdef' for c in argb):
                if len(argb) == 8:
                    return "#" + argb[2:] # Strip alpha
                return "#" + argb
        
        # Fallback ou gestion future des thèmes (index)
        # if hasattr(color_obj, 'theme') ... (nécessiterait de charger le thème du workbook)
        return None
    except:
        return None

def extract_excel_data():
    print("--- STARTING EXTRACTION V3 (Fix Colors) ---")
    print(f"Chargement du fichier : {EXCEL_FILE}")
    if not os.path.exists(EXCEL_FILE):
        print(f"Erreur : Le fichier {EXCEL_FILE} est introuvable.")
        return

    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True) # data_only=True pour avoir les valeurs calculées
        
        all_data = {}
        
        for sheet_name in wb.sheetnames:
            print(f"Traitement de la feuille : {sheet_name}")
            ws = wb[sheet_name]
            
            # 1. Identifier les colonnes visibles
            visible_columns_indices = [] # 1-based indices
            
            # Openpyxl column_dimensions keys are letters 'A', 'B'...
            # Iterons sur les colonnes utilisées
            max_col = ws.max_column
            max_row = min(ws.max_row, 100) # Limite pour perfs prototype
            
            visible_headers = []
            
            # On suppose que la ligne header est quelque part au début.
            # Pour faire simple avec les styles, on va prendre tout le bloc et le frontend triera.
            # Mais l'utilisateur veut exclure les colonnes masquées.
            
            # Map index (1-based) -> is_hidden
            hidden_cols_map = {}
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                if col_letter in ws.column_dimensions:
                    cd = ws.column_dimensions[col_letter]
                    if cd.hidden:
                        hidden_cols_map[col_idx] = True
            
            # Extraction des données ligne par ligne
            rows_data = []
            
            # Chercher le 1er header visible (simple heuristique : ligne avec du texte)
            header_row_idx = 1
            for r in range(1, 20):
                if ws.row_dimensions[r].hidden:
                    continue
                # Check content
                row_values = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1) if not hidden_cols_map.get(c)]
                non_empty = sum(1 for v in row_values if v and isinstance(v, str))
                if non_empty > 2:
                    header_row_idx = r
                    visible_headers = [str(v) if v else "" for v in row_values]
                    break
            
            if not visible_headers:
                 # Fallback
                 visible_headers = [f"Col {i}" for i in range(1, max_col + 1) if not hidden_cols_map.get(i)]

            # Lire les données
            for r in range(header_row_idx + 1, max_row + 1):
                if ws.row_dimensions[r].hidden:
                    continue
                
                row_obj = {}
                col_counter = 0
                
                # Iteration colonnes
                for c in range(1, max_col + 1):
                    if hidden_cols_map.get(c):
                        continue
                        
                    cell = ws.cell(row=r, column=c)
                    val = cell.value
                    
                    # Style extraction
                    bg_color = get_rgb_color(cell.fill.fgColor) if cell.fill else None
                    font_color = get_rgb_color(cell.font.color) if cell.font else None
                    is_bold = cell.font.b if cell.font else False
                    is_italic = cell.font.i if cell.font else False
                    
                    # Formatage particulier
                    display_val = val
                    if isinstance(val, float):
                        display_val = f"{val:.2f}"
                    elif val is None:
                        display_val = ""
                        
                    # Structure Cellule Riche
                    cell_data = {
                        "value": display_val,
                        "style": {}
                    }
                    if bg_color and bg_color != "#000000" and bg_color != "#FFFFFF": # Ignore black/white default sometimes logic
                         # Check if transparency/default
                         if not (bg_color.startswith("#00") and len(bg_color) == 9):
                            cell_data["style"]["backgroundColor"] = bg_color
                    
                    if font_color and font_color != "#000000": # Ignore default black
                        cell_data["style"]["color"] = font_color

                    if is_bold:
                        cell_data["style"]["fontWeight"] = "bold"
                    
                    if is_italic:
                        cell_data["style"]["fontStyle"] = "italic"
                    
                    # Clé colonne correspondante au header
                    # Si plus de colonnes de données que de headers, on ignore ou on gère
                    if col_counter < len(visible_headers):
                        col_name = visible_headers[col_counter]
                        # Gestion des clés dupliquées dans le header
                        if col_name in row_obj:
                            col_name = f"{col_name}_{col_counter}"
                        row_obj[col_name] = cell_data
                        
                    col_counter += 1
                
                rows_data.append(row_obj)

            sheet_data = {
                "columns": visible_headers,
                "rows": rows_data,
                "rowCount": len(rows_data)
            }
            all_data[sheet_name] = sheet_data

        js_content = f"const EXCEL_DATA = {json.dumps(all_data, indent=4, ensure_ascii=False, cls=DateTimeEncoder)};"
        
        with open(OUTPUT_JS_FILE, "w", encoding="utf-8") as f:
            f.write(js_content)
            
        print(f"Succès ! Données exportées avec styles dans {OUTPUT_JS_FILE}")

    except Exception as e:
        print(f"Erreur : {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    extract_excel_data()

